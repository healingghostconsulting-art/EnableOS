from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
import csv

OUTPUT_DIR = Path('/home/ubuntu/chcg-enableos-demo/references')
XLSX_PATH = OUTPUT_DIR / 'CHCG_EnableOS_Development_Plan.xlsx'
CSV_PATH = OUTPUT_DIR / 'CHCG_EnableOS_Development_Plan.csv'
SUMMARY_PATH = OUTPUT_DIR / 'CHCG_EnableOS_Development_Plan_Summary.md'

rows = [
    {
        'Area': 'Completed Foundation',
        'Item': 'Core full-stack application foundation established with React 19, Tailwind 4, Express, tRPC, Drizzle, tenant-aware data model, and Manus OAuth integration.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'The platform now has a stable demo-ready technical base that can support continued feature expansion.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Role-based workspace framework delivered for learner, coach, manager, executive, admin, reporting, and content-facing experiences.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Primary navigation, route access, and role-specific shells are in place.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Mission-control UI shell created with collapsible sidebar, compact headers, shared visual language, and lower-scroll workspace patterns.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'This provides the design system direction, but a broader consistency pass is still recommended.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Learner training experience converted from stacked scroll-heavy sections into guided page-by-page progression with contextual navigation.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Improves orientation and reduces visual fatigue during training.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Interactive flash-card brief system redesigned across Training Zone and Content Missions, including composition cleanup and jump-strip navigation.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Brief content is more navigable and visually cleaner than the previous fragmented layout.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Coach workspace refined with visible coaching-log launch actions and a focused reusable pop-up flow for structured log entry.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'The coach ribbon interaction was restored and validated after a visibility regression.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Manager workspace expanded with a shared coaching-log modal path so coaching actions can occur without leaving the current context.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Supports role continuity and reuse of shared coaching components.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Content Missions redesigned into a cleaner training-library style experience with focused launch behavior for selected learning content.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Reduces clutter and creates a more intentional content-entry experience.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Content Missions explorer regression repaired so track cards, filters, and text hierarchy render clearly again.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'This closed a visible presentation issue that undermined client demos.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Reporting workspace established with executive-facing proof, ROI, error-rate, question, lifecycle, cohort, and benchmarking views.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Creates a strong narrative for client-facing analytics demonstrations.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Executive dashboard strengthened with tenure-aware lifecycle reporting, proof-of-impact framing, coaching consistency, and behavior-analysis insights.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'These are persuasive demo surfaces, though live production data integration remains future work.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Question-level readiness and performance intelligence added, including miss-rate trends, first-pass success, and high-alert risk states.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Supports evidence-led conversations about training effectiveness.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Learner journey and assigned re-engagement flows refined with stronger wording, improved readability, and better transition behavior.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Includes terminology updates and completion-state improvements.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Training quiz and knowledge-check interfaces reformatted for cleaner spacing, stronger readability, and a more modern instructional feel.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Important groundwork for curriculum expansion, but more instructional-content depth is still needed.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Role-scoped navigation and access behavior tightened so learner experiences are better isolated from coach, manager, admin, and executive areas.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Helps preserve a more credible multi-role demo path.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Training brief completion, coaching-log save, and related learner/coach regressions repaired and validated.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Stabilized several high-visibility flows that had previously blocked demos.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Targeted regression coverage expanded across training layout, flash-card briefs, coaching popups, reporting surfaces, and content-library behavior.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Automated validation supports safer future UI iterations.'
    },
    {
        'Area': 'Completed Foundation',
        'Item': 'Homepage and launch surfaces compacted with denser proof-led messaging inspired by modern security-training product patterns.',
        'Status': 'Completed',
        'Priority': 'Completed',
        'Est. Hours': 'Delivered',
        'Notes': 'Creates a stronger first impression, though brand and messaging can continue to mature.'
    },
    {
        'Area': 'In Progress / Current Sprint',
        'Item': 'Finalize the client-facing roadmap package and align the product narrative around what is demo-ready versus what remains to reach production depth.',
        'Status': 'In Progress',
        'Priority': 'High',
        'Est. Hours': '4-6 hrs',
        'Notes': 'This spreadsheet is part of that packaging work and should become the editable client planning artifact.'
    },
    {
        'Area': 'In Progress / Current Sprint',
        'Item': 'Continue refinements to the new Content Missions library experience so launch behavior, readability, and hierarchy remain stable as more curriculum is added.',
        'Status': 'In Progress',
        'Priority': 'High',
        'Est. Hours': '8-16 hrs',
        'Notes': 'The main redesign is complete, but the surface will need another pass when more real content is migrated.'
    },
    {
        'Area': 'In Progress / Current Sprint',
        'Item': 'Continue polish on the page-based training flow so lesson transitions, progress states, and in-course content density feel more intentional across all modules.',
        'Status': 'In Progress',
        'Priority': 'High',
        'Est. Hours': '12-24 hrs',
        'Notes': 'This is especially important once the full seven-module curriculum is loaded in production depth.'
    },
    {
        'Area': 'In Progress / Current Sprint',
        'Item': 'Monitor the reusable coaching-log popup pattern across coach and manager flows and apply any follow-up adjustments found during stakeholder review.',
        'Status': 'In Progress',
        'Priority': 'Medium',
        'Est. Hours': '6-10 hrs',
        'Notes': 'The core workflow works, but client review may surface spacing or form-usability adjustments.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Migrate the full training curriculum into the LMS at production depth across all seven planned modules, including learner and leadership tracks.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '120-180 hrs',
        'Notes': 'This is the single largest remaining body of work and likely the most important dependency for a production-ready rollout.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Create a structured content-ingestion and maintenance workflow so curriculum updates can be managed without repeated manual UI assembly.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '24-40 hrs',
        'Notes': 'Recommended before or during full curriculum migration to prevent future bottlenecks.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Run a product-wide UI consistency pass across every role workspace, modal, card system, typography scale, spacing pattern, and state transition.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '40-60 hrs',
        'Notes': 'The product is significantly improved, but it still needs a broader unification pass to feel fully coherent end to end.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Complete mobile and tablet responsiveness for priority workflows so the LMS is credible across common device sizes.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '32-48 hrs',
        'Notes': 'Current work has been optimized primarily for desktop demo quality rather than full responsive production readiness.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Replace demo-style assumptions with production-grade authentication, SSO, tenant administration, and enterprise access-management controls.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '32-56 hrs',
        'Notes': 'Critical for real enterprise deployment and stakeholder trust.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Connect reporting experiences to live operational data pipelines and define trustworthy source-of-truth metrics for coaching, QA, and learning outcomes.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '40-80 hrs',
        'Notes': 'Current analytics storytelling is strong for demonstration, but production reporting needs validated live data integration.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Build exportable reporting outputs for client stakeholders, including spreadsheet, PDF, presentation, or email-ready summaries where appropriate.',
        'Status': 'Queued',
        'Priority': 'Medium',
        'Est. Hours': '16-28 hrs',
        'Notes': 'Important for operational adoption and client communications.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Implement a notification and reminder system for coaching actions, assignment nudges, escalations, and manager follow-through.',
        'Status': 'Queued',
        'Priority': 'Medium',
        'Est. Hours': '24-40 hrs',
        'Notes': 'Would materially improve day-to-day usefulness beyond demo interactions.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Create admin-facing content and program management tools so non-developers can manage modules, briefs, assignments, and supporting artifacts.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '36-60 hrs',
        'Notes': 'This reduces long-term dependence on engineering for routine curriculum changes.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Deepen coaching workflow functionality with edit history, approvals, linked action plans, attachment support, and richer cadence management.',
        'Status': 'Queued',
        'Priority': 'Medium',
        'Est. Hours': '24-36 hrs',
        'Notes': 'The current workflow proves the interaction model, but operational maturity will need more depth.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Expand learner training states with stronger progress persistence, assignment orchestration, recommendation logic, and completion governance.',
        'Status': 'Queued',
        'Priority': 'Medium',
        'Est. Hours': '24-40 hrs',
        'Notes': 'Important once the full curriculum and real user traffic are in place.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Run a formal accessibility review covering color contrast, keyboard navigation, focus management, semantic structure, and assistive-technology behavior.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '20-32 hrs',
        'Notes': 'Recommended before external rollout so remediation is built into the next polish cycle.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Perform performance optimization for large single-page views, training media loads, and state-heavy dashboards as real content volume increases.',
        'Status': 'Queued',
        'Priority': 'Medium',
        'Est. Hours': '18-30 hrs',
        'Notes': 'Especially relevant because the main UI currently lives in a very large view file.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Refactor oversized UI and router files into more maintainable feature modules to improve delivery speed and reduce regression risk.',
        'Status': 'Queued',
        'Priority': 'Medium',
        'Est. Hours': '28-44 hrs',
        'Notes': 'This is a technical-health investment that will make future roadmap delivery safer.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Add stronger QA coverage for role-specific end-to-end scenarios, reporting accuracy checks, and curriculum-content regression verification.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '24-40 hrs',
        'Notes': 'Useful before introducing a larger body of curriculum and more production-oriented integrations.'
    },
    {
        'Area': 'Queued / Roadmap',
        'Item': 'Prepare production hardening work across environment configuration, error monitoring, data backup practices, auditability, and release readiness.',
        'Status': 'Queued',
        'Priority': 'High',
        'Est. Hours': '30-50 hrs',
        'Notes': 'Necessary for moving from persuasive demo environment to a dependable production program.'
    },
]

summary_metrics = [
    ('Current position', 'Strong demo-quality foundation with significant remaining work to reach production depth.', 'The product already demonstrates multi-role workflows, guided training, coaching, content exploration, and reporting.'),
    ('Major completed milestone groups', '18 grouped foundations delivered', 'Completed items are intentionally summarized as milestone themes rather than hundreds of individual tasks.'),
    ('Current sprint focus', 'Packaging, training-flow polish, content-library stabilization, and coaching modal follow-up', 'These items keep the current demo stable while clarifying next-stage priorities.'),
    ('Largest remaining workstream', 'Full curriculum migration into all 7 modules', 'This is the most substantial remaining effort across learner and leadership training tracks.'),
    ('Additional high-priority gaps', 'UI consistency, mobile responsiveness, enterprise auth/SSO, live data integration, admin tools, accessibility', 'These items define the difference between a compelling demo and a production-ready platform.'),
    ('Planning note', 'All estimates are directional planning ranges, not fixed implementation commitments.', 'Actual effort will depend on curriculum source quality, stakeholder revision cycles, and production integration requirements.'),
]

notes = [
    'The platform now has a persuasive, client-demo-ready foundation across learner, coach, manager, executive, admin, reporting, and content-library experiences.',
    'At the same time, there is still substantial work remaining in three major tracks: UI/UX consistency and polish, functional depth for real operations, and migration of the full CHCG training curriculum into the LMS.',
    'Recommended sequencing is to complete the curriculum/content plan, then run a unified UI and responsiveness pass, then finalize production-grade operational features such as SSO, live reporting integrations, exports, notifications, and admin management tooling.',
]

status_fill = {
    'Completed': PatternFill('solid', fgColor='D9EAD3'),
    'In Progress': PatternFill('solid', fgColor='FFF2CC'),
    'Queued': PatternFill('solid', fgColor='F4CCCC'),
}

priority_fill = {
    'High': PatternFill('solid', fgColor='F4CCCC'),
    'Medium': PatternFill('solid', fgColor='FCE5CD'),
    'Low': PatternFill('solid', fgColor='D9EAD3'),
    'Completed': PatternFill('solid', fgColor='D9EAD3'),
}

thin = Side(style='thin', color='D0D7DE')
header_fill = PatternFill('solid', fgColor='1F4E78')
header_font = Font(color='FFFFFF', bold=True)
title_font = Font(size=15, bold=True)
section_font = Font(size=12, bold=True)
wrap = Alignment(vertical='top', wrap_text=True)

wb = Workbook()
ws_summary = wb.active
ws_summary.title = 'Summary'
ws_plan = wb.create_sheet('Development Plan')
ws_timeline = wb.create_sheet('Roadmap Totals')

# Summary sheet
ws_summary['A1'] = 'CHCG EnableOS Development Plan Summary'
ws_summary['A1'].font = title_font
ws_summary['A3'] = 'Planning Context'
ws_summary['A3'].font = section_font
for i, note in enumerate(notes, start=4):
    ws_summary[f'A{i}'] = note
    ws_summary[f'A{i}'].alignment = wrap

summary_header_row = 9
for col, title in enumerate(['Metric', 'Value', 'Notes'], start=1):
    cell = ws_summary.cell(row=summary_header_row, column=col, value=title)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = wrap
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_index, metric in enumerate(summary_metrics, start=summary_header_row + 1):
    for col_index, value in enumerate(metric, start=1):
        cell = ws_summary.cell(row=row_index, column=col_index, value=value)
        cell.alignment = wrap
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws_summary.column_dimensions['A'].width = 28
ws_summary.column_dimensions['B'].width = 42
ws_summary.column_dimensions['C'].width = 55

# Plan sheet
headers = ['Area', 'Item', 'Status', 'Priority', 'Est. Hours', 'Notes']
for col, header in enumerate(headers, start=1):
    cell = ws_plan.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = wrap
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, row in enumerate(rows, start=2):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_plan.cell(row=row_idx, column=col_idx, value=row[header])
        cell.alignment = wrap
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if header == 'Status' and row[header] in status_fill:
            cell.fill = status_fill[row[header]]
        if header == 'Priority' and row[header] in priority_fill:
            cell.fill = priority_fill[row[header]]

widths = {
    'A': 24,
    'B': 92,
    'C': 14,
    'D': 14,
    'E': 16,
    'F': 58,
}
for col, width in widths.items():
    ws_plan.column_dimensions[col].width = width

ws_plan.freeze_panes = 'A2'
last_row = len(rows) + 1
tab = Table(displayName='DevelopmentPlanTable', ref=f'A1:F{last_row}')
style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws_plan.add_table(tab)

# Timeline totals sheet
ws_timeline['A1'] = 'Roadmap Effort Snapshot'
ws_timeline['A1'].font = title_font
for col, title in enumerate(['Category', 'Approx. Open Items', 'Combined Estimate', 'Commentary'], start=1):
    cell = ws_timeline.cell(row=3, column=col, value=title)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = wrap
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

roadmap_totals = [
    ('Current sprint follow-up', 4, '30-56 hrs', 'Covers packaging, Content Missions stabilization, training-flow polish, and popup workflow follow-up.'),
    ('Curriculum and content migration', 2, '144-220 hrs', 'Includes full seven-module curriculum migration plus a sustainable content-ingestion workflow.'),
    ('Experience polish and responsiveness', 2, '72-108 hrs', 'Includes product-wide UI consistency work and mobile/tablet responsiveness.'),
    ('Enterprise functionality and operations', 7, '186-330 hrs', 'Includes SSO/auth depth, live data integration, exports, notifications, admin tools, learner/workflow depth, and production hardening.'),
    ('Quality, accessibility, and maintainability', 4, '90-146 hrs', 'Includes accessibility review, performance optimization, refactoring, and expanded QA coverage.'),
]

for row_idx, values in enumerate(roadmap_totals, start=4):
    for col_idx, value in enumerate(values, start=1):
        cell = ws_timeline.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = wrap
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws_timeline['A11'] = 'Overall planning range'
ws_timeline['B11'] = '19 open roadmap items'
ws_timeline['C11'] = '522-860 hrs'
ws_timeline['D11'] = 'Directional estimate only. The total will move based on curriculum-source readiness, stakeholder reviews, and production integration scope.'
for col in range(1, 5):
    cell = ws_timeline.cell(row=11, column=col)
    cell.font = Font(bold=True)
    cell.alignment = wrap
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws_timeline.column_dimensions['A'].width = 30
ws_timeline.column_dimensions['B'].width = 20
ws_timeline.column_dimensions['C'].width = 18
ws_timeline.column_dimensions['D'].width = 70

wb.save(XLSX_PATH)

with CSV_PATH.open('w', newline='', encoding='utf-8') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=headers)
    writer.writeheader()
    writer.writerows(rows)

summary_md = f'''# CHCG EnableOS Development Plan Summary

## Current Position

CHCG EnableOS now has a **strong demo-quality foundation** across learner, coach, manager, executive, admin, reporting, and content-library experiences. The platform already demonstrates role-based navigation, guided training flows, coaching workflows, content discovery, and executive storytelling.

At the same time, the roadmap should be framed honestly: there is still **substantial work remaining** before the platform is ready for a production-grade client rollout. The largest remaining gaps fall into three main tracks:

| Major track | Current reality | Why it matters |
| --- | --- | --- |
| UI and UX polish | The product is much stronger visually than earlier versions, but it still needs a full consistency, responsiveness, and accessibility pass. | This determines whether the experience feels unified and credible across all workspaces. |
| Functional depth | Core flows exist, but several enterprise features still need to move from demo-ready to operations-ready. | This includes SSO, live data integrations, exports, notifications, admin tooling, and production hardening. |
| Curriculum and content migration | Only part of the learning experience is represented at full depth today. | The full seven-module CHCG curriculum still needs to be imposed into the LMS in a scalable, maintainable way. |

## What the Attached Workbook Includes

The attached workbook contains three editable sheets:

| Sheet | Purpose |
| --- | --- |
| Summary | A plain-language overview of the current state, major gaps, and planning assumptions. |
| Development Plan | A client-facing roadmap with grouped completed milestones, current sprint items, queued work, priority, estimated time, and notes. |
| Roadmap Totals | A simple effort roll-up by workstream so the client can understand the scale of the remaining effort. |

## Planning Note

All time ranges in the workbook are **directional planning estimates**, not fixed delivery commitments. Actual effort will depend on curriculum-source readiness, stakeholder revision cycles, and the degree of production integration required.
'''
SUMMARY_PATH.write_text(summary_md, encoding='utf-8')

print(f'Created: {XLSX_PATH}')
print(f'Created: {CSV_PATH}')
print(f'Created: {SUMMARY_PATH}')
